"""
Views для отчётов и аналитики
"""

import json
from rest_framework import generics, status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse, FileResponse
from django.utils import timezone
from django.db.models import Q
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter

from .models import ReportTemplate, GeneratedReport, Dashboard, AnalyticsQuery
from .serializers import (
    ReportTemplateSerializer, GeneratedReportSerializer, GenerateReportSerializer,
    DashboardSerializer, AnalyticsQuerySerializer, ExecuteQuerySerializer,
    ProjectAnalyticsSerializer, UserPerformanceSerializer, SystemAnalyticsSerializer,
    ExportDataSerializer, ChartDataSerializer
)
from .services import ReportGenerator, AnalyticsService
from apps.projects.models import Project
from apps.defects.models import Defect
from apps.common.permissions import IsProjectMember
from celery import current_app


class ReportTemplateListCreateView(generics.ListCreateAPIView):
    """
    Список шаблонов отчётов и создание нового шаблона
    """
    serializer_class = ReportTemplateSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['name', 'description']
    filterset_fields = ['report_type', 'output_format', 'is_active']
    ordering = ['name']
    
    def get_queryset(self):
        """Получение доступных шаблонов"""
        user = self.request.user
        
        # Администраторы видят все шаблоны
        if user.is_admin:
            return ReportTemplate.objects.all()
        
        # Остальные видят публичные шаблоны и свои собственные
        return ReportTemplate.objects.filter(
            Q(is_public=True) | Q(created_by=user)
        ).filter(is_active=True)
    
    def perform_create(self, serializer):
        """Создание шаблона"""
        # Проверяем права на создание шаблонов
        user = self.request.user
        if not (user.is_admin or user.role in ['manager']):
            raise ValidationError("Недостаточно прав для создания шаблонов отчётов")
        
        serializer.save()


class ReportTemplateDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Детали, обновление и удаление шаблона отчёта
    """
    serializer_class = ReportTemplateSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Получение доступных шаблонов"""
        user = self.request.user
        
        if user.is_admin:
            return ReportTemplate.objects.all()
        
        return ReportTemplate.objects.filter(
            Q(is_public=True) | Q(created_by=user)
        )
    
    def get_permissions(self):
        """Права доступа"""
        if self.request.method in ['PUT', 'PATCH', 'DELETE']:
            # Изменять могут только создатели или администраторы
            self.permission_classes = [permissions.IsAuthenticated]
            # Дополнительная проверка в методах
        return super().get_permissions()
    
    def perform_update(self, serializer):
        """Обновление шаблона"""
        template = self.get_object()
        user = self.request.user
        
        if not (template.created_by == user or user.is_admin):
            raise ValidationError("Недостаточно прав для изменения шаблона")
        
        serializer.save()
    
    def perform_destroy(self, instance):
        """Удаление шаблона"""
        user = self.request.user
        
        if not (instance.created_by == user or user.is_admin):
            raise ValidationError("Недостаточно прав для удаления шаблона")
        
        # Мягкое удаление
        instance.is_active = False
        instance.save()


class GeneratedReportListView(generics.ListAPIView):
    """
    Список сгенерированных отчётов
    """
    serializer_class = GeneratedReportSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['name', 'description']
    filterset_fields = ['status', 'template', 'project']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Получение отчётов пользователя"""
        user = self.request.user
        
        # Администраторы видят все отчёты
        if user.is_admin:
            return GeneratedReport.objects.all()
        
        # Остальные видят только свои отчёты
        return GeneratedReport.objects.filter(generated_by=user)


class GeneratedReportDetailView(generics.RetrieveDestroyAPIView):
    """
    Детали и удаление сгенерированного отчёта
    """
    serializer_class = GeneratedReportSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Получение отчётов пользователя"""
        user = self.request.user
        
        if user.is_admin:
            return GeneratedReport.objects.all()
        
        return GeneratedReport.objects.filter(generated_by=user)


class GenerateReportView(APIView):
    """
    Генерация нового отчёта
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        """Запуск генерации отчёта"""
        serializer = GenerateReportSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(
                serializer.errors,
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Проверяем доступ к шаблону
        template = serializer.validated_data['template']
        user = request.user
        
        if not template.is_public and template.created_by != user and not user.is_admin:
            return Response(
                {"error": "Нет доступа к этому шаблону"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Проверяем доступ к проекту
        project = serializer.validated_data.get('project')
        if project and not (project.is_member(user) or user.is_admin):
            return Response(
                {"error": "Нет доступа к этому проекту"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Создаём запись отчёта
        report = GeneratedReport.objects.create(
            template=template,
            name=serializer.validated_data['name'],
            description=serializer.validated_data.get('description', ''),
            project=project,
            date_from=serializer.validated_data.get('date_from'),
            date_to=serializer.validated_data.get('date_to'),
            filter_params=serializer.validated_data.get('filter_params', {}),
            generated_by=user,
            status=GeneratedReport.Status.PENDING
        )
        
        # Запускаем генерацию в фоне (если есть Celery) или синхронно
        try:
            # Попробуем запустить через Celery
            from .tasks import generate_report_task
            generate_report_task.delay(report.id)
        except ImportError:
            # Если Celery не настроен, генерируем синхронно
            generator = ReportGenerator(report)
            generator.generate()
        
        return Response(
            GeneratedReportSerializer(report, context={'request': request}).data,
            status=status.HTTP_201_CREATED
        )


class DownloadReportView(APIView):
    """
    Скачивание сгенерированного отчёта
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, report_id):
        """Скачивание файла отчёта"""
        try:
            user = request.user
            
            # Получаем отчёт
            if user.is_admin:
                report = GeneratedReport.objects.get(id=report_id)
            else:
                report = GeneratedReport.objects.get(id=report_id, generated_by=user)
            
            # Проверяем статус
            if report.status != GeneratedReport.Status.COMPLETED:
                return Response(
                    {"error": "Отчёт ещё не готов"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Проверяем срок действия
            if report.is_expired:
                return Response(
                    {"error": "Срок действия отчёта истёк"},
                    status=status.HTTP_410_GONE
                )
            
            # Проверяем наличие файла
            if not report.file:
                return Response(
                    {"error": "Файл отчёта не найден"},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Отмечаем скачивание
            report.mark_downloaded()
            
            # Возвращаем файл
            return FileResponse(
                report.file.open('rb'),
                as_attachment=True,
                filename=report.file.name.split('/')[-1]
            )
            
        except GeneratedReport.DoesNotExist:
            return Response(
                {"error": "Отчёт не найден"},
                status=status.HTTP_404_NOT_FOUND
            )


class DashboardListCreateView(generics.ListCreateAPIView):
    """
    Список дашбордов и создание нового дашборда
    """
    serializer_class = DashboardSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'description']
    ordering = ['dashboard_type', 'name']
    
    def get_queryset(self):
        """Получение доступных дашбордов"""
        user = self.request.user
        
        # Администраторы видят все дашборды
        if user.is_admin:
            return Dashboard.objects.all()
        
        # Остальные видят публичные дашборды, свои и доступные по роли
        return Dashboard.objects.filter(
            Q(is_public=True) |
            Q(created_by=user) |
            Q(allowed_roles__contains=[user.role])
        ).distinct()


class DashboardDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Детали, обновление и удаление дашборда
    """
    serializer_class = DashboardSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Получение доступных дашбордов"""
        user = self.request.user
        
        if user.is_admin:
            return Dashboard.objects.all()
        
        return Dashboard.objects.filter(
            Q(is_public=True) |
            Q(created_by=user) |
            Q(allowed_roles__contains=[user.role])
        ).distinct()


class AnalyticsQueryListCreateView(generics.ListCreateAPIView):
    """
    Список аналитических запросов и создание нового запроса
    """
    serializer_class = AnalyticsQuerySerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['name', 'description']
    filterset_fields = ['query_type', 'is_public']
    ordering = ['query_type', 'name']
    
    def get_queryset(self):
        """Получение доступных запросов"""
        user = self.request.user
        
        if user.is_admin:
            return AnalyticsQuery.objects.all()
        
        return AnalyticsQuery.objects.filter(
            Q(is_public=True) | Q(created_by=user)
        )


class AnalyticsQueryDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Детали, обновление и удаление аналитического запроса
    """
    serializer_class = AnalyticsQuerySerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Получение доступных запросов"""
        user = self.request.user
        
        if user.is_admin:
            return AnalyticsQuery.objects.all()
        
        return AnalyticsQuery.objects.filter(
            Q(is_public=True) | Q(created_by=user)
        )


class ExecuteQueryView(APIView):
    """
    Выполнение аналитического запроса
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        """Выполнение запроса"""
        serializer = ExecuteQuerySerializer(
            data=request.data,
            context={'request': request}
        )
        
        if not serializer.is_valid():
            return Response(
                serializer.errors,
                status=status.HTTP_400_BAD_REQUEST
            )
        
        query = serializer.validated_data['query']
        parameters = serializer.validated_data.get('parameters', {})
        
        try:
            # Выполняем SQL запрос (упрощённая реализация)
            from django.db import connection
            
            with connection.cursor() as cursor:
                cursor.execute(query.sql_query, parameters)
                columns = [col[0] for col in cursor.description]
                results = [
                    dict(zip(columns, row))
                    for row in cursor.fetchall()
                ]
            
            # Отмечаем использование запроса
            query.mark_used()
            
            return Response({
                'query_name': query.name,
                'columns': columns,
                'results': results,
                'count': len(results)
            })
            
        except Exception as e:
            return Response(
                {"error": f"Ошибка выполнения запроса: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def project_analytics(request, project_id):
    """
    Аналитика по конкретному проекту
    """
    try:
        project = Project.objects.get(id=project_id)
        
        # Проверяем доступ к проекту
        user = request.user
        if not (project.is_member(user) or user.is_admin):
            return Response(
                {"error": "Нет доступа к этому проекту"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Получаем параметры периода
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        if date_from:
            date_from = timezone.datetime.fromisoformat(date_from)
        if date_to:
            date_to = timezone.datetime.fromisoformat(date_to)
        
        # Получаем аналитику
        analytics = AnalyticsService.get_project_analytics(
            project, date_from, date_to
        )
        
        serializer = ProjectAnalyticsSerializer(analytics)
        return Response(serializer.data)
        
    except Project.DoesNotExist:
        return Response(
            {"error": "Проект не найден"},
            status=status.HTTP_404_NOT_FOUND
        )
    except ValueError as e:
        return Response(
            {"error": f"Ошибка параметров: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def user_performance(request, user_id=None):
    """
    Статистика производительности пользователя
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    # Если ID не указан, возвращаем статистику текущего пользователя
    if not user_id:
        target_user = request.user
    else:
        try:
            target_user = User.objects.get(id=user_id)
            
            # Проверяем права доступа
            user = request.user
            if not (user.is_admin or user.role == 'manager' or user == target_user):
                return Response(
                    {"error": "Недостаточно прав"},
                    status=status.HTTP_403_FORBIDDEN
                )
                
        except User.DoesNotExist:
            return Response(
                {"error": "Пользователь не найден"},
                status=status.HTTP_404_NOT_FOUND
            )
    
    # Получаем параметры периода
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        date_from = timezone.datetime.fromisoformat(date_from)
    if date_to:
        date_to = timezone.datetime.fromisoformat(date_to)
    
    # Получаем статистику
    performance = AnalyticsService.get_user_performance(
        target_user, date_from, date_to
    )
    
    serializer = UserPerformanceSerializer(performance)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def system_analytics(request):
    """
    Общая аналитика системы
    """
    # Проверяем права доступа
    user = request.user
    if not (user.is_admin or user.role == 'manager'):
        return Response(
            {"error": "Недостаточно прав"},
            status=status.HTTP_403_FORBIDDEN
        )
    
    # Получаем аналитику
    analytics = AnalyticsService.get_system_analytics()
    
    serializer = SystemAnalyticsSerializer(analytics)
    return Response(serializer.data)


class ExportDataView(APIView):
    """
    Экспорт данных в различных форматах
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        """Экспорт данных"""
        serializer = ExportDataSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(
                serializer.errors,
                status=status.HTTP_400_BAD_REQUEST
            )
        
        data_type = serializer.validated_data['data_type']
        export_format = serializer.validated_data['export_format']
        filters = serializer.validated_data.get('filters', {})
        fields = serializer.validated_data.get('fields', [])
        
        try:
            # Получаем данные для экспорта
            if data_type == 'defects':
                data = self._get_defects_data(request.user, filters, fields)
                filename = f"defects_export.{export_format}"
            elif data_type == 'projects':
                data = self._get_projects_data(request.user, filters, fields)
                filename = f"projects_export.{export_format}"
            else:
                return Response(
                    {"error": "Неподдерживаемый тип данных"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Генерируем файл
            if export_format == 'csv':
                content = self._generate_csv(data)
                content_type = 'text/csv'
            elif export_format == 'excel':
                content = self._generate_excel(data)
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif export_format == 'json':
                content = json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')
                content_type = 'application/json'
            else:
                return Response(
                    {"error": "Неподдерживаемый формат"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Возвращаем файл
            response = HttpResponse(content, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            return Response(
                {"error": f"Ошибка экспорта: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _get_defects_data(self, user, filters, fields):
        """Получение данных дефектов для экспорта"""
        # Базовый queryset с учётом прав доступа
        if user.is_admin:
            defects = Defect.objects.all()
        else:
            user_projects = user.projects.all()
            managed_projects = user.managed_projects.all() if user.is_manager else []
            all_projects = user_projects.union(managed_projects) if managed_projects else user_projects
            defects = Defect.objects.filter(project__in=all_projects)
        
        # Применяем фильтры
        for key, value in filters.items():
            if key == 'project' and value:
                defects = defects.filter(project_id=value)
            elif key == 'status' and value:
                defects = defects.filter(status__in=value)
            elif key == 'priority' and value:
                defects = defects.filter(priority__in=value)
        
        # Выбираем поля для экспорта
        if not fields:
            fields = [
                'defect_number', 'title', 'description', 'status', 'priority',
                'project__name', 'category__name', 'author__first_name',
                'author__last_name', 'assignee__first_name', 'assignee__last_name',
                'created_at', 'due_date'
            ]
        
        return list(defects.values(*fields))
    
    def _get_projects_data(self, user, filters, fields):
        """Получение данных проектов для экспорта"""
        # Базовый queryset с учётом прав доступа
        if user.is_admin:
            projects = Project.objects.all()
        else:
            user_projects = user.projects.all()
            managed_projects = user.managed_projects.all() if user.is_manager else []
            projects = user_projects.union(managed_projects) if managed_projects else user_projects
        
        # Выбираем поля для экспорта
        if not fields:
            fields = [
                'name', 'description', 'status', 'priority',
                'manager__first_name', 'manager__last_name',
                'start_date', 'end_date', 'created_at'
            ]
        
        return list(projects.values(*fields))
    
    def _generate_csv(self, data):
        """Генерация CSV"""
        import csv
        import io
        
        output = io.StringIO()
        
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        content = output.getvalue()
        output.close()
        
        return content.encode('utf-8-sig')
    
    def _generate_excel(self, data):
        """Генерация Excel"""
        from openpyxl import Workbook
        import io
        
        wb = Workbook()
        ws = wb.active
        
        if data:
            # Заголовки
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Данные
            for row, item in enumerate(data, 2):
                for col, value in enumerate(item.values(), 1):
                    ws.cell(row=row, column=col, value=value)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def chart_data(request):
    """
    Получение данных для графиков
    """
    chart_type = request.GET.get('type', 'defects_by_status')
    project_id = request.GET.get('project_id')
    
    user = request.user
    
    try:
        if chart_type == 'defects_by_status':
            # График дефектов по статусам
            defects_qs = Defect.objects.all()
            
            if project_id:
                project = Project.objects.get(id=project_id)
                if not (project.is_member(user) or user.is_admin):
                    return Response(
                        {"error": "Нет доступа к проекту"},
                        status=status.HTTP_403_FORBIDDEN
                    )
                defects_qs = defects_qs.filter(project=project)
            elif not user.is_admin:
                # Ограничиваем доступные проекты
                user_projects = user.projects.all()
                managed_projects = user.managed_projects.all() if user.is_manager else []
                all_projects = user_projects.union(managed_projects) if managed_projects else user_projects
                defects_qs = defects_qs.filter(project__in=all_projects)
            
            data = dict(
                defects_qs.values('status').annotate(
                    count=Count('id')
                ).values_list('status', 'count')
            )
            
            chart_data = {
                'chart_type': 'pie',
                'title': 'Дефекты по статусам',
                'data': {
                    'labels': list(data.keys()),
                    'datasets': [{
                        'data': list(data.values()),
                        'backgroundColor': [
                            '#ff6384', '#36a2eb', '#ffce56', '#4bc0c0', '#9966ff'
                        ]
                    }]
                }
            }
            
        elif chart_type == 'defects_trend':
            # График динамики дефектов
            from datetime import timedelta
            
            end_date = timezone.now()
            start_date = end_date - timedelta(days=30)
            
            defects_qs = Defect.objects.filter(
                created_at__range=[start_date, end_date]
            )
            
            if project_id:
                project = Project.objects.get(id=project_id)
                if not (project.is_member(user) or user.is_admin):
                    return Response(
                        {"error": "Нет доступа к проекту"},
                        status=status.HTTP_403_FORBIDDEN
                    )
                defects_qs = defects_qs.filter(project=project)
            
            # Группируем по дням
            dates = []
            counts = []
            current_date = start_date.date()
            
            while current_date <= end_date.date():
                count = defects_qs.filter(created_at__date=current_date).count()
                dates.append(current_date.strftime('%d.%m'))
                counts.append(count)
                current_date += timedelta(days=1)
            
            chart_data = {
                'chart_type': 'line',
                'title': 'Динамика создания дефектов',
                'data': {
                    'labels': dates,
                    'datasets': [{
                        'label': 'Количество дефектов',
                        'data': counts,
                        'borderColor': '#36a2eb',
                        'fill': False
                    }]
                }
            }
            
        else:
            return Response(
                {"error": "Неподдерживаемый тип графика"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = ChartDataSerializer(chart_data)
        return Response(serializer.data)
        
    except Project.DoesNotExist:
        return Response(
            {"error": "Проект не найден"},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {"error": f"Ошибка получения данных: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
