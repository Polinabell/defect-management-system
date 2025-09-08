"""
Сервисы для генерации отчётов и аналитики
"""

import io
import csv
import json
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from django.db.models import Count, Avg, Q, F
from django.utils import timezone
from django.core.files.base import ContentFile
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from apps.projects.models import Project, ProjectStage
from apps.defects.models import Defect, DefectCategory, DefectComment
from .models import GeneratedReport, AnalyticsQuery

User = get_user_model()


class ReportGenerator:
    """
    Сервис для генерации отчётов
    """
    
    def __init__(self, report: GeneratedReport):
        self.report = report
        self.template = report.template
        self.project = report.project
        self.date_from = report.date_from
        self.date_to = report.date_to
        self.filter_params = report.filter_params
    
    def generate(self) -> bool:
        """
        Основной метод генерации отчёта
        """
        try:
            self.report.status = GeneratedReport.Status.PROCESSING
            self.report.save()
            
            start_time = timezone.now()
            
            # Получаем данные в зависимости от типа отчёта
            data = self._get_report_data()
            
            # Генерируем файл в зависимости от формата
            file_content = self._generate_file(data)
            
            # Сохраняем файл
            filename = self._generate_filename()
            self.report.file.save(
                filename,
                ContentFile(file_content),
                save=False
            )
            
            # Обновляем статус
            self.report.status = GeneratedReport.Status.COMPLETED
            self.report.generated_at = timezone.now()
            self.report.processing_time = self.report.generated_at - start_time
            self.report.file_size = len(file_content)
            self.report.expires_at = timezone.now() + timedelta(days=30)  # Срок действия 30 дней
            self.report.save()
            
            return True
            
        except Exception as e:
            self.report.status = GeneratedReport.Status.FAILED
            self.report.error_message = str(e)
            self.report.save()
            return False
    
    def _get_report_data(self) -> Dict[str, Any]:
        """
        Получение данных для отчёта в зависимости от типа
        """
        if self.template.report_type == 'project_summary':
            return self._get_project_summary_data()
        elif self.template.report_type == 'defects_analysis':
            return self._get_defects_analysis_data()
        elif self.template.report_type == 'performance_report':
            return self._get_performance_report_data()
        elif self.template.report_type == 'timeline_report':
            return self._get_timeline_report_data()
        else:
            return self._get_custom_report_data()
    
    def _get_project_summary_data(self) -> Dict[str, Any]:
        """Данные для сводки по проекту"""
        project = self.project
        if not project:
            raise ValueError("Для отчёта по проекту необходимо указать проект")
        
        # Базовая информация о проекте
        data = {
            'project': {
                'name': project.name,
                'description': project.description,
                'manager': project.manager.get_full_name(),
                'start_date': project.start_date,
                'end_date': project.end_date,
                'status': project.get_status_display(),
                'progress': project.progress_percentage,
            }
        }
        
        # Дефекты проекта
        defects_qs = project.defects.all()
        if self.date_from:
            defects_qs = defects_qs.filter(created_at__date__gte=self.date_from)
        if self.date_to:
            defects_qs = defects_qs.filter(created_at__date__lte=self.date_to)
        
        data['defects'] = {
            'total': defects_qs.count(),
            'by_status': dict(defects_qs.values('status').annotate(count=Count('id')).values_list('status', 'count')),
            'by_priority': dict(defects_qs.values('priority').annotate(count=Count('id')).values_list('priority', 'count')),
            'by_category': dict(defects_qs.values('category__name').annotate(count=Count('id')).values_list('category__name', 'count')),
            'overdue': defects_qs.filter(due_date__lt=timezone.now().date(), status__in=['new', 'in_progress']).count(),
        }
        
        # Этапы проекта
        stages = project.stages.all()
        data['stages'] = []
        for stage in stages:
            data['stages'].append({
                'name': stage.name,
                'status': stage.get_status_display(),
                'progress': stage.completion_percentage,
                'start_date': stage.start_date,
                'end_date': stage.end_date,
                'defects_count': stage.defects.count(),
            })
        
        # Участники проекта
        members = project.project_members.filter(is_active=True)
        data['members'] = []
        for member in members:
            assigned_defects = defects_qs.filter(assignee=member.user)
            data['members'].append({
                'name': member.user.get_full_name(),
                'role': member.get_role_display(),
                'assigned_defects': assigned_defects.count(),
                'closed_defects': assigned_defects.filter(status='closed').count(),
            })
        
        return data
    
    def _get_defects_analysis_data(self) -> Dict[str, Any]:
        """Данные для анализа дефектов"""
        # Базовый queryset дефектов
        defects_qs = Defect.objects.all()
        
        if self.project:
            defects_qs = defects_qs.filter(project=self.project)
        
        if self.date_from:
            defects_qs = defects_qs.filter(created_at__date__gte=self.date_from)
        
        if self.date_to:
            defects_qs = defects_qs.filter(created_at__date__lte=self.date_to)
        
        # Применяем дополнительные фильтры
        for key, value in self.filter_params.items():
            if key == 'status' and value:
                defects_qs = defects_qs.filter(status__in=value)
            elif key == 'priority' and value:
                defects_qs = defects_qs.filter(priority__in=value)
            elif key == 'category' and value:
                defects_qs = defects_qs.filter(category__in=value)
        
        data = {
            'summary': {
                'total_defects': defects_qs.count(),
                'open_defects': defects_qs.exclude(status__in=['closed', 'cancelled']).count(),
                'closed_defects': defects_qs.filter(status='closed').count(),
                'overdue_defects': defects_qs.filter(
                    due_date__lt=timezone.now().date(),
                    status__in=['new', 'in_progress']
                ).count(),
            }
        }
        
        # Статистика по статусам
        data['by_status'] = dict(
            defects_qs.values('status').annotate(count=Count('id')).values_list('status', 'count')
        )
        
        # Статистика по приоритетам
        data['by_priority'] = dict(
            defects_qs.values('priority').annotate(count=Count('id')).values_list('priority', 'count')
        )
        
        # Статистика по категориям
        data['by_category'] = dict(
            defects_qs.values('category__name').annotate(count=Count('id')).values_list('category__name', 'count')
        )
        
        # Статистика по серьёзности
        data['by_severity'] = dict(
            defects_qs.values('severity').annotate(count=Count('id')).values_list('severity', 'count')
        )
        
        # Топ исполнителей
        data['top_assignees'] = list(
            defects_qs.exclude(assignee__isnull=True)
            .values('assignee__first_name', 'assignee__last_name')
            .annotate(
                count=Count('id'),
                closed_count=Count('id', filter=Q(status='closed'))
            )
            .order_by('-count')[:10]
        )
        
        # Среднее время решения
        closed_defects = defects_qs.filter(status='closed', closed_at__isnull=False)
        if closed_defects.exists():
            resolution_times = []
            for defect in closed_defects:
                if defect.resolution_time:
                    resolution_times.append(defect.resolution_time)
            
            if resolution_times:
                data['average_resolution_time'] = sum(resolution_times) / len(resolution_times)
            else:
                data['average_resolution_time'] = 0
        else:
            data['average_resolution_time'] = 0
        
        return data
    
    def _get_performance_report_data(self) -> Dict[str, Any]:
        """Данные для отчёта о производительности"""
        # Период анализа
        if not self.date_from:
            self.date_from = timezone.now().date() - timedelta(days=30)
        if not self.date_to:
            self.date_to = timezone.now().date()
        
        data = {'period': {'from': self.date_from, 'to': self.date_to}}
        
        # Получаем пользователей для анализа
        users_qs = User.objects.filter(is_active=True)
        if self.project:
            users_qs = users_qs.filter(
                Q(projects=self.project) | Q(managed_projects=self.project)
            ).distinct()
        
        user_performance = []
        
        for user in users_qs:
            # Дефекты пользователя за период
            created_defects = Defect.objects.filter(
                author=user,
                created_at__date__range=[self.date_from, self.date_to]
            )
            
            assigned_defects = Defect.objects.filter(
                assignee=user,
                assigned_at__date__range=[self.date_from, self.date_to]
            )
            
            closed_defects = Defect.objects.filter(
                assignee=user,
                status='closed',
                closed_at__date__range=[self.date_from, self.date_to]
            )
            
            # Комментарии пользователя
            comments_count = DefectComment.objects.filter(
                author=user,
                created_at__date__range=[self.date_from, self.date_to]
            ).count()
            
            # Среднее время решения
            resolution_times = []
            for defect in closed_defects:
                if defect.resolution_time:
                    resolution_times.append(defect.resolution_time)
            
            avg_resolution_time = sum(resolution_times) / len(resolution_times) if resolution_times else 0
            
            user_performance.append({
                'user': user.get_full_name(),
                'role': user.get_role_display(),
                'defects_created': created_defects.count(),
                'defects_assigned': assigned_defects.count(),
                'defects_closed': closed_defects.count(),
                'comments_count': comments_count,
                'avg_resolution_time': avg_resolution_time,
                'overdue_defects': assigned_defects.filter(
                    due_date__lt=timezone.now().date(),
                    status__in=['new', 'in_progress']
                ).count(),
            })
        
        data['user_performance'] = sorted(
            user_performance,
            key=lambda x: x['defects_closed'],
            reverse=True
        )
        
        return data
    
    def _get_timeline_report_data(self) -> Dict[str, Any]:
        """Данные для временного отчёта"""
        if not self.date_from:
            self.date_from = timezone.now().date() - timedelta(days=30)
        if not self.date_to:
            self.date_to = timezone.now().date()
        
        data = {'period': {'from': self.date_from, 'to': self.date_to}}
        
        # Дефекты за период
        defects_qs = Defect.objects.filter(
            created_at__date__range=[self.date_from, self.date_to]
        )
        
        if self.project:
            defects_qs = defects_qs.filter(project=self.project)
        
        # Группируем по дням
        timeline_data = []
        current_date = self.date_from
        
        while current_date <= self.date_to:
            day_defects = defects_qs.filter(created_at__date=current_date)
            closed_defects = Defect.objects.filter(
                closed_at__date=current_date
            )
            
            if self.project:
                closed_defects = closed_defects.filter(project=self.project)
            
            timeline_data.append({
                'date': current_date,
                'defects_created': day_defects.count(),
                'defects_closed': closed_defects.count(),
                'critical_defects': day_defects.filter(priority='critical').count(),
                'high_priority_defects': day_defects.filter(priority='high').count(),
            })
            
            current_date += timedelta(days=1)
        
        data['timeline'] = timeline_data
        
        return data
    
    def _get_custom_report_data(self) -> Dict[str, Any]:
        """Данные для пользовательского отчёта"""
        # Здесь можно реализовать логику для пользовательских отчётов
        # на основе конфигурации в template.filter_config
        return {'message': 'Пользовательские отчёты в разработке'}
    
    def _generate_file(self, data: Dict[str, Any]) -> bytes:
        """Генерация файла отчёта"""
        if self.template.output_format == 'csv':
            return self._generate_csv(data)
        elif self.template.output_format == 'excel':
            return self._generate_excel(data)
        elif self.template.output_format == 'json':
            return self._generate_json(data)
        else:
            return self._generate_pdf(data)
    
    def _generate_csv(self, data: Dict[str, Any]) -> bytes:
        """Генерация CSV файла"""
        output = io.StringIO()
        
        if self.template.report_type == 'defects_analysis':
            writer = csv.writer(output)
            
            # Заголовки
            writer.writerow(['Метрика', 'Значение'])
            
            # Общая статистика
            writer.writerow(['Общее количество дефектов', data['summary']['total_defects']])
            writer.writerow(['Открытые дефекты', data['summary']['open_defects']])
            writer.writerow(['Закрытые дефекты', data['summary']['closed_defects']])
            writer.writerow(['Просроченные дефекты', data['summary']['overdue_defects']])
            
            # Статистика по статусам
            writer.writerow([])
            writer.writerow(['Статус', 'Количество'])
            for status, count in data['by_status'].items():
                writer.writerow([status, count])
        
        elif self.template.report_type == 'performance_report':
            writer = csv.writer(output)
            
            # Заголовки
            writer.writerow([
                'Пользователь', 'Роль', 'Создано дефектов', 'Назначено дефектов',
                'Закрыто дефектов', 'Комментариев', 'Среднее время решения (ч)',
                'Просроченные дефекты'
            ])
            
            # Данные пользователей
            for user_data in data['user_performance']:
                writer.writerow([
                    user_data['user'],
                    user_data['role'],
                    user_data['defects_created'],
                    user_data['defects_assigned'],
                    user_data['defects_closed'],
                    user_data['comments_count'],
                    round(user_data['avg_resolution_time'], 2),
                    user_data['overdue_defects'],
                ])
        
        content = output.getvalue()
        output.close()
        
        return content.encode('utf-8-sig')  # BOM для корректного отображения в Excel
    
    def _generate_excel(self, data: Dict[str, Any]) -> bytes:
        """Генерация Excel файла"""
        wb = Workbook()
        ws = wb.active
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if self.template.report_type == 'project_summary':
            ws.title = "Сводка по проекту"
            
            # Заголовок
            ws['A1'] = f"Сводка по проекту: {data['project']['name']}"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:D1')
            
            # Информация о проекте
            row = 3
            ws[f'A{row}'] = "Название проекта"
            ws[f'B{row}'] = data['project']['name']
            row += 1
            
            ws[f'A{row}'] = "Менеджер"
            ws[f'B{row}'] = data['project']['manager']
            row += 1
            
            ws[f'A{row}'] = "Статус"
            ws[f'B{row}'] = data['project']['status']
            row += 1
            
            ws[f'A{row}'] = "Прогресс"
            ws[f'B{row}'] = f"{data['project']['progress']}%"
            row += 2
            
            # Дефекты
            ws[f'A{row}'] = "Статистика дефектов"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws[f'A{row}'] = "Всего дефектов"
            ws[f'B{row}'] = data['defects']['total']
            row += 1
            
            ws[f'A{row}'] = "Просроченных"
            ws[f'B{row}'] = data['defects']['overdue']
            row += 2
            
            # Статистика по статусам
            ws[f'A{row}'] = "Статус"
            ws[f'B{row}'] = "Количество"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'].font = header_font
            ws[f'B{row}'].fill = header_fill
            row += 1
            
            for status, count in data['defects']['by_status'].items():
                ws[f'A{row}'] = status
                ws[f'B{row}'] = count
                row += 1
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение в байты
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _generate_json(self, data: Dict[str, Any]) -> bytes:
        """Генерация JSON файла"""
        # Сериализация datetime объектов
        def json_serializer(obj):
            if isinstance(obj, (datetime, timezone.datetime)):
                return obj.isoformat()
            elif hasattr(obj, 'isoformat'):
                return obj.isoformat()
            raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
        
        json_data = {
            'report': {
                'name': self.report.name,
                'template': self.template.name,
                'generated_at': timezone.now().isoformat(),
                'period': {
                    'from': self.date_from.isoformat() if self.date_from else None,
                    'to': self.date_to.isoformat() if self.date_to else None,
                }
            },
            'data': data
        }
        
        return json.dumps(json_data, ensure_ascii=False, indent=2, default=json_serializer).encode('utf-8')
    
    def _generate_pdf(self, data: Dict[str, Any]) -> bytes:
        """Генерация PDF файла (заглушка)"""
        # Здесь можно реализовать генерацию PDF с помощью reportlab или weasyprint
        # Пока возвращаем JSON как замену
        return self._generate_json(data)
    
    def _generate_filename(self) -> str:
        """Генерация имени файла"""
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        extension = self.template.output_format
        if extension == 'excel':
            extension = 'xlsx'
        
        return f"{self.template.name}_{timestamp}.{extension}"


class AnalyticsService:
    """
    Сервис для аналитики и статистики
    """
    
    @staticmethod
    def get_project_analytics(project: Project, date_from: Optional[datetime] = None, 
                            date_to: Optional[datetime] = None) -> Dict[str, Any]:
        """Получение аналитики по проекту"""
        # Дефекты проекта
        defects_qs = project.defects.all()
        
        if date_from:
            defects_qs = defects_qs.filter(created_at__gte=date_from)
        if date_to:
            defects_qs = defects_qs.filter(created_at__lte=date_to)
        
        # Базовая статистика
        total_defects = defects_qs.count()
        
        # Статистика по статусам
        defects_by_status = dict(
            defects_qs.values('status').annotate(count=Count('id')).values_list('status', 'count')
        )
        
        # Статистика по приоритетам
        defects_by_priority = dict(
            defects_qs.values('priority').annotate(count=Count('id')).values_list('priority', 'count')
        )
        
        # Статистика по категориям
        defects_by_category = dict(
            defects_qs.values('category__name').annotate(count=Count('id')).values_list('category__name', 'count')
        )
        
        # Временная аналитика (дефекты по дням за последние 30 дней)
        end_date = date_to or timezone.now()
        start_date = date_from or (end_date - timedelta(days=30))
        
        defects_by_day = []
        current_date = start_date.date()
        
        while current_date <= end_date.date():
            day_created = defects_qs.filter(created_at__date=current_date).count()
            day_closed = defects_qs.filter(closed_at__date=current_date).count()
            
            defects_by_day.append({
                'date': current_date.isoformat(),
                'created': day_created,
                'closed': day_closed,
            })
            
            current_date += timedelta(days=1)
        
        # Производительность
        closed_defects = defects_qs.filter(status='closed', closed_at__isnull=False)
        
        if closed_defects.exists():
            resolution_times = []
            for defect in closed_defects:
                if defect.resolution_time:
                    resolution_times.append(defect.resolution_time)
            
            avg_resolution_time = sum(resolution_times) / len(resolution_times) if resolution_times else 0
        else:
            avg_resolution_time = 0
        
        # Просроченные дефекты
        overdue_defects = defects_qs.filter(
            due_date__lt=timezone.now().date(),
            status__in=['new', 'in_progress']
        ).count()
        
        # Дефекты по исполнителям
        defects_by_assignee = dict(
            defects_qs.exclude(assignee__isnull=True)
            .values('assignee__first_name', 'assignee__last_name')
            .annotate(count=Count('id'))
            .values_list('assignee__first_name', 'count')  # Упрощённо, только имя
        )
        
        return {
            'total_defects': total_defects,
            'defects_by_status': defects_by_status,
            'defects_by_priority': defects_by_priority,
            'defects_by_category': defects_by_category,
            'defects_created_by_day': defects_by_day,
            'average_resolution_time': avg_resolution_time,
            'overdue_defects': overdue_defects,
            'defects_by_assignee': defects_by_assignee,
            'project_progress': project.progress_percentage,
        }
    
    @staticmethod
    def get_user_performance(user: User, date_from: Optional[datetime] = None,
                           date_to: Optional[datetime] = None) -> Dict[str, Any]:
        """Получение статистики производительности пользователя"""
        # Фильтр по датам
        date_filter = {}
        if date_from:
            date_filter['created_at__gte'] = date_from
        if date_to:
            date_filter['created_at__lte'] = date_to
        
        # Дефекты, созданные пользователем
        defects_created = Defect.objects.filter(author=user, **date_filter).count()
        
        # Дефекты, назначенные пользователю
        assigned_filter = date_filter.copy()
        if date_from:
            assigned_filter['assigned_at__gte'] = date_from
        if date_to:
            assigned_filter['assigned_at__lte'] = date_to
        
        defects_assigned = Defect.objects.filter(assignee=user, **assigned_filter).count()
        
        # Дефекты, закрытые пользователем
        closed_filter = date_filter.copy()
        if date_from:
            closed_filter['closed_at__gte'] = date_from
        if date_to:
            closed_filter['closed_at__lte'] = date_to
        
        defects_closed = Defect.objects.filter(
            assignee=user, 
            status='closed',
            **closed_filter
        ).count()
        
        # Среднее время решения
        closed_defects = Defect.objects.filter(
            assignee=user,
            status='closed',
            closed_at__isnull=False
        )
        
        resolution_times = []
        for defect in closed_defects:
            if defect.resolution_time:
                resolution_times.append(defect.resolution_time)
        
        avg_resolution_time = sum(resolution_times) / len(resolution_times) if resolution_times else 0
        
        # Просроченные дефекты
        overdue_defects = Defect.objects.filter(
            assignee=user,
            due_date__lt=timezone.now().date(),
            status__in=['new', 'in_progress']
        ).count()
        
        # Комментарии
        comment_filter = date_filter.copy()
        comments_count = DefectComment.objects.filter(author=user, **comment_filter).count()
        
        return {
            'user_id': user.id,
            'user_name': user.get_full_name(),
            'defects_created': defects_created,
            'defects_assigned': defects_assigned,
            'defects_closed': defects_closed,
            'average_resolution_time': avg_resolution_time,
            'overdue_defects': overdue_defects,
            'comments_count': comments_count,
            'last_activity': user.last_login,
        }
    
    @staticmethod
    def get_system_analytics() -> Dict[str, Any]:
        """Получение общей аналитики системы"""
        # Общие показатели
        total_projects = Project.objects.count()
        total_defects = Defect.objects.count()
        total_users = User.objects.count()
        active_users = User.objects.filter(is_active=True).count()
        
        # Статистика дефектов
        defects_by_status = dict(
            Defect.objects.values('status').annotate(count=Count('id')).values_list('status', 'count')
        )
        
        defects_by_priority = dict(
            Defect.objects.values('priority').annotate(count=Count('id')).values_list('priority', 'count')
        )
        
        # Дефекты за этот месяц
        current_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        defects_created_this_month = Defect.objects.filter(created_at__gte=current_month).count()
        defects_closed_this_month = Defect.objects.filter(
            status='closed',
            closed_at__gte=current_month
        ).count()
        
        # Статистика проектов
        projects_by_status = dict(
            Project.objects.values('status').annotate(count=Count('id')).values_list('status', 'count')
        )
        
        overdue_projects = Project.objects.filter(
            end_date__lt=timezone.now().date(),
            status__in=['planning', 'in_progress']
        ).count()
        
        completed_projects_this_month = Project.objects.filter(
            status='completed',
            updated_at__gte=current_month
        ).count()
        
        # Среднее время решения дефектов
        closed_defects = Defect.objects.filter(status='closed', closed_at__isnull=False)
        resolution_times = []
        
        for defect in closed_defects:
            if defect.resolution_time:
                resolution_times.append(defect.resolution_time)
        
        avg_resolution_time = sum(resolution_times) / len(resolution_times) if resolution_times else 0
        
        return {
            'total_projects': total_projects,
            'total_defects': total_defects,
            'total_users': total_users,
            'active_users': active_users,
            'defects_by_status': defects_by_status,
            'defects_by_priority': defects_by_priority,
            'defects_created_this_month': defects_created_this_month,
            'defects_closed_this_month': defects_closed_this_month,
            'projects_by_status': projects_by_status,
            'overdue_projects': overdue_projects,
            'completed_projects_this_month': completed_projects_this_month,
            'average_resolution_time': avg_resolution_time,
        }
